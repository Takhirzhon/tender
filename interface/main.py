import re
import streamlit as st
from datetime import datetime, timedelta
import json
import os
import sys
import time
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from io import BytesIO
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from core.score_matrix import AVK5Estimator, DocumentComplianceChecker, ProfitabilityAnalyzer
import requests
import xml.etree.ElementTree as ET
from urllib.parse import urlencode

PROZORRO_API_URL = "https://public.api.openprocurement.org/api/2.4/tenders"
OUTPUT_DIR = "../tenders"
MAX_RESULTS = 3
RATE_LIMIT_DELAY = 1.5
# Load environment variables
load_dotenv()

def setup_environment():
    """Create output directory if it doesn't exist"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"📁 Output directory created: {OUTPUT_DIR}")
def download_prozorro_tenders(topic=None, total_to_download=1, days_back=None):
    """
    Download tenders from ProZorro API based on tender topic (using keywords.json)
    """
    print(f"🔍 Downloading tenders for topic: {topic}")

    # Load topic keywords
    keywords_path = os.path.join("data", "keywords.json")
    if not os.path.exists(keywords_path):
        raise FileNotFoundError("❌ keywords.json not found!")

    with open(keywords_path, "r", encoding="utf-8") as f:
        topic_keywords = json.load(f).get(topic, [])
    
    if not topic_keywords:
        raise ValueError(f"❌ No keywords found for topic '{topic}' in keywords.json")

    setup_environment()
    
    downloaded = []
    offset_time = datetime.now()
    checked = 0

    while len(downloaded) < total_to_download and checked < 500:
        params = {
            "descending": 1,
            "limit": 100,
            "offset": offset_time.strftime("%Y-%m-%dT%H:%M:%S")
        }

        try:
            response = requests.get(PROZORRO_API_URL, params=params)
            response.raise_for_status()
            tenders = response.json().get("data", [])
            if not tenders:
                print("🚫 No more tenders found.")
                break

            for tender in tenders:
                tender_id = tender["id"]
                try:
                    tender_url = f"{PROZORRO_API_URL}/{tender_id}"
                    tender_response = requests.get(tender_url)
                    tender_response.raise_for_status()
                    tender_data = tender_response.json()["data"]

                    title = tender_data.get("title", "").lower()
                    description = tender_data.get("description", "").lower()

                    # Check if any keyword matches
                    if any(kw.lower() in title or kw.lower() in description for kw in topic_keywords):
                        filename = f"ProZorro_{tender_id}.json"
                        filepath = os.path.join(OUTPUT_DIR, filename)

                        with open(filepath, "w", encoding="utf-8") as f:
                            json.dump(tender_data, f, ensure_ascii=False, indent=2)

                        downloaded.append({
                            "id": tender_id,
                            "title": tender_data.get("title", "Без назви"),
                            "date": tender_data.get("dateModified", ""),
                            "budget": tender_data.get("value", {}).get("amount", 0),
                            "file": filename
                        })
                        print(f"✅ Saved: {filename}")

                        if len(downloaded) >= total_to_download:
                            break

                    checked += 1
                    time.sleep(RATE_LIMIT_DELAY)

                except Exception as e:
                    print(f"⚠️ Error for {tender_id}: {e}")
                    continue

            # Update offset to last tender timestamp
            offset_time -= timedelta(minutes=10)

        except Exception as e:
            print(f"❌ API error: {e}")
            break

    print(f"\n💾 Total downloaded tenders: {len(downloaded)} for topic: {topic}")
    return downloaded

# Initialize Claude client
def get_claude_client():
    api_key = os.getenv("CLAUDE_API_KEY")
    if not api_key:
        st.error("❌ Claude API key not found in .env file")
        return None
    return anthropic.Anthropic(api_key=api_key)

# Enhanced parser function
def analyze_tender(text, client):
    prompt = f"""
You are an expert in Ukrainian public procurement tenders. Analyze the tender text and extract the following information:

1. Basic Information:
   - Title or Project Name
   - Issuer or Client
   - Submission Deadline
   - Estimated Budget (with currency)
   - Location (City, Region)
   - Project Type/Scope

2. Critical Requirements:
   - List ALL required documents (comma-separated)
   - Does this tender require PC AVK5 cost estimates? (true/false)
   - Key technical specifications (summarize key requirements)

3. Financial & Legal:
   - Payment terms and schedule
   - References to Ukrainian laws/regulations (list)

4. Viability Analysis:
   - Resource requirements (equipment, personnel, etc.)
   - Timeline feasibility assessment (adequate/risky/inadequate)
   - Profitability assessment (high/medium/low)

Return the result STRICTLY in JSON format with these keys:
{{
  "title": "...",
  "issuer": "...",
  "deadline": "...",
  "budget": "...",
  "location": "...",
  "project_type": "...",
  "required_documents": ["doc1", "doc2", ...],
  "avk5_required": true/false,
  "technical_specs": "...",
  "payment_terms": "...",
  "resource_requirements": "...",
  "timeline_feasibility": "...",
  "profitability": "..."
}}

Tender Text:
\"\"\"
{text[:15000]}  # Truncate to 15,000 characters
\"\"\"
"""
    try:
        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=1024,
            temperature=0.0,
            system="You are a procurement specialist analyzing Ukrainian tenders. Focus on PC AVK5 compliance and document requirements.",
            messages=[{"role": "user", "content": prompt}]
        )
        
        if response.content:
            result = ''.join(block.text for block in response.content if hasattr(block, 'text'))
            try:
                return json.loads(result)
            except json.JSONDecodeError:
                match = re.search(r'({.*})', result, re.DOTALL)
                if match:
                    try:
                        return json.loads(match.group(1))
                    except:
                        pass
                st.error("❌ Claude returned invalid JSON.")
                return {}
    except Exception as e:
        st.error(f"❌ Error analyzing tender: {e}")
        return {}

# Excel formatting
def format_excel(ws):
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    col_widths = [40, 30, 15, 15, 20, 25, 40, 15, 50, 30, 40, 20, 20, 30]
    columns = [
        "Title", "Issuer", "Deadline", "Budget", "Location", 
        "Project Type", "Required Documents", "PC AVK5 Required",
        "Technical Specifications", "Payment Terms", "Resource Requirements",
        "Timeline Feasibility", "Profitability Assessment", "Filename"
    ]
    
    for col_num, (column_title, width) in enumerate(zip(columns, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    return ws, columns

# Streamlit App
st.set_page_config(page_title="AI Tender Optimizer", layout="wide")

# Initialize session state
if "analysis_results" not in st.session_state:
    st.session_state.analysis_results = []
if "tenders_downloaded" not in st.session_state:
    st.session_state.tenders_downloaded = []
if "excel_buffer" not in st.session_state:
    st.session_state.excel_buffer = None
if "analysis_attempted" not in st.session_state:
    st.session_state.analysis_attempted = False
if "company_resources" not in st.session_state:
    st.session_state.company_resources = {
        "workers": 10,
        "engineers": 2,
        "vehicles": 3,
        "current_projects": []
    }
if "document_vault" not in st.session_state:
    st.session_state.document_vault = DocumentComplianceChecker()

# Sidebar
st.sidebar.title("🛠️ Tender Settings")
tab = st.sidebar.radio("Navigation", ["📥 Data Downloader", "🔍 Tender Analysis", "🏢 Company Profile", "📊 Tender Evaluation"])

keywords_option = st.sidebar.multiselect(
    "🔍 Select Tender Topics",
    ["IT", "Construction", "Medical Equipment", "Transport", "Education"],
    default=["Construction"]
)

days_range = st.sidebar.slider(
    "📅 Days Back to Search",
    min_value=1,
    max_value=60,
    value=30
)

# Main Page
st.title("📦 AI Tender Optimizer")

if tab == "📥 Data Downloader":
    st.header("📥 Download Tenders from ProZorro")

    # Load topics from keywords.json dynamically
    with open("/mount/src/tender/interface/data/keywords.json", "r", encoding="utf-8") as f:
        topic_keywords = json.load(f)
    topic_list = list(topic_keywords.keys())

    topic = st.selectbox("📚 Choose Tender Topic", topic_list)
    num_tenders = st.slider("📦 Number of Tenders to Download", min_value=5, max_value=100, value=20, step=5)
    days_back = st.slider("📅 Search Tenders from Last N Days", min_value=1, max_value=60, value=30)

    if st.button("🚀 Start Download"):
        with st.spinner("Downloading..."):
            tenders = download_prozorro_tenders(
                topic=topic,
                total_to_download=num_tenders,
                days_back=days_back
            )

        if tenders:
            st.success(f"✅ {len(tenders)} tenders downloaded for topic '{topic}'")
            st.session_state["tenders_downloaded"] = tenders

            with st.expander("📄 Tender Summary"):
                st.json({
                    "topic": topic,
                    "keywords": topic_keywords[topic],
                    "total_downloaded": len(tenders)
                })

            st.download_button(
                label="📁 Download Tender Metadata",
                data=json.dumps(tenders, ensure_ascii=False, indent=2),
                file_name=f"{topic.lower()}_tenders_summary.json",
                mime="application/json"
            )
        else:
            st.warning("❌ No tenders matched your criteria.")


elif tab == "🔍 Tender Analysis":
    st.header("🔍 Tender Analysis")

    # Check for existing tender files
    def load_existing_tenders():
        tender_files = []
        tenders_dir = "../tenders"
        if os.path.exists(tenders_dir):
            for filename in os.listdir(tenders_dir):
                if filename.startswith("ProZorro_") and filename.endswith(".json"):
                    tender_id = filename.replace("ProZorro_", "").replace(".json", "")
                    try:
                        with open(os.path.join(tenders_dir, filename), "r", encoding="utf-8") as f:
                            tender_data = json.load(f)
                            tender_files.append({
                                "id": tender_id,
                                "title": tender_data.get("title", "Без назви"),
                                "date": tender_data.get("dateModified", ""),
                                "budget": tender_data.get("value", {}).get("amount", 0),
                                "file": filename
                            })
                    except Exception as e:
                        st.warning(f"Error loading {filename}: {e}")
        return tender_files

    # Load tenders if missing
    if not st.session_state.tenders_downloaded:
        existing = load_existing_tenders()
        if existing:
            st.session_state.tenders_downloaded = existing
            st.success(f"✅ Loaded {len(existing)} tenders from folder.")
        else:
            st.warning("⚠️ No tender files found.")
            st.stop()

    tender_options = {t['id']: t['title'] for t in st.session_state.tenders_downloaded}
    selected_tenders = st.multiselect(
        "Select tenders to analyze:",
        options=list(tender_options.keys()),
        format_func=lambda x: f"{x} - {tender_options[x][:50]}..."
    )

    if not selected_tenders:
        st.info("ℹ️ Please select at least one tender to analyze.")
        st.stop()

    client = get_claude_client()
    if not client:
        st.stop()

    status_text = st.empty()
    analyze_clicked = st.button("🔍 Analyze Selected Tenders", key="analyze_button")

    if analyze_clicked:
        st.session_state.analysis_attempted = True
        results = []
        progress_bar = st.progress(0)

        for i, tid in enumerate(selected_tenders):
            tender = next((t for t in st.session_state.tenders_downloaded if t['id'] == tid), None)
            if not tender:
                continue

            progress_bar.progress((i + 1) / len(selected_tenders))
            status_text.text(f"Analyzing {tid}...")

            path = f"../tenders/ProZorro_{tid}.json"
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                content = f"""
Tender Title: {data.get('title', '')}
Issuer: {data.get('procuringEntity', {}).get('name', '')}
Location: {data.get('procuringEntity', {}).get('address', {}).get('locality', '')}, {data.get('procuringEntity', {}).get('address', {}).get('region', '')}
Budget: {data.get('value', {}).get('amount', '')} {data.get('value', {}).get('currency', 'UAH')}
Deadline: {data.get('tenderPeriod', {}).get('endDate', '')}
Description: {data.get('description', '')}
""".strip()
                result = analyze_tender(content, client)
                if result:
                    result["tender_id"] = tid
                    result["Filename"] = f"{tid}.txt"
                    results.append(result)
            time.sleep(1.5)

        st.session_state.analysis_results = results
        status_text.text("✅ Analysis complete.")
        progress_bar.empty()

        # Save to Excel
        if results:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet("Tender Analysis")
            else:
                ws.title = "Tender Analysis"
            ws, columns = format_excel(ws)
            for idx, res in enumerate(results, 2):
                row_data = [
                    res.get("title", "N/A"),
                    res.get("issuer", "N/A"),
                    res.get("deadline", "N/A"),
                    res.get("budget", "N/A"),
                    res.get("location", "N/A"),
                    res.get("project_type", "N/A"),
                    ", ".join(res.get("required_documents", [])),
                    "Yes" if res.get("avk5_required", False) else "No",
                    res.get("technical_specs", "N/A"),
                    res.get("payment_terms", "N/A"),
                    res.get("resource_requirements", "N/A"),
                    res.get("timeline_feasibility", "N/A"),
                    res.get("profitability", "N/A"),
                    res.get("Filename", "N/A")
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.session_state.excel_buffer = buf

    # 🔁 Always restore results
    results = st.session_state.get("analysis_results", [])
    if not results:
        st.warning("⚠️ No analysis results found. Run analysis first.")
        st.stop()

    # 📋 SHOW ANALYSIS RESULTS
    st.subheader("Analysis Results")
    for res in results:
        with st.expander(f"{res.get('title', 'Untitled')} - {res.get('tender_id', '')}"):
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Issuer", res.get("issuer", "N/A"))
                st.metric("Deadline", res.get("deadline", "N/A"))
                st.metric("Budget", res.get("budget", "N/A"))
                st.metric("Location", res.get("location", "N/A"))
                st.metric("Project Type", res.get("project_type", "N/A"))
                st.metric("PC AVK5 Required", "✅ Yes" if res.get("avk5_required") else "❌ No")
            with col2:
                st.subheader("Required Documents")
                for doc in res.get("required_documents", []):
                    st.write(f"- {doc}")
                st.subheader("Technical Specifications")
                st.info(res.get("technical_specs", "No technical specs"))
                st.subheader("Viability")
                st.metric("Timeline Feasibility", res.get("timeline_feasibility", "N/A"))
                st.metric("Profitability", res.get("profitability", "N/A"))

    if st.session_state.excel_buffer:
        st.download_button(
            label="📊 Download Full Analysis (Excel)",
            data=st.session_state.excel_buffer,
            file_name="tender_analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="excel_download_button"
        )

elif tab == "🏢 Company Profile":
    st.header("🏢 Company Profile Management")
    
    st.subheader("Company Resources")
    col1, col2, col3 = st.columns(3)
    workers = col1.number_input("Number of Workers", 
                               value=st.session_state.company_resources["workers"],
                               min_value=0, step=1)
    engineers = col2.number_input("Number of Engineers", 
                                 value=st.session_state.company_resources["engineers"],
                                 min_value=0, step=1)
    vehicles = col3.number_input("Number of Vehicles", 
                                value=st.session_state.company_resources["vehicles"],
                                min_value=0, step=1)
    
    # Current projects
    st.subheader("Current Projects")
    projects = st.session_state.company_resources["current_projects"]
    for i, project in enumerate(projects):
        col1, col2, col3 = st.columns([3, 2, 1])
        project_name = col1.text_input(f"Project {i+1} Name", value=project["name"])
        duration = col2.number_input(f"Duration (days)", value=project["duration"], min_value=1)
        if col3.button("❌", key=f"del_proj_{i}"):
            projects.pop(i)
            st.rerun()
    
    if st.button("➕ Add Project"):
        projects.append({"name": "New Project", "duration": 30})
        st.rerun()
    
    # Document vault
    st.subheader("Document Vault")
    st.write("Upload company documents for compliance checking")
    
    doc_name = st.text_input("Document Name")
    doc_type = st.text_input("Document Type")
    validity = st.date_input("Validity Date")
    uploaded_file = st.file_uploader("Upload Document")
    
    if st.button("Add to Vault") and uploaded_file:
        # Save file
        os.makedirs("data/documents", exist_ok=True)
        file_path = f"data/documents/{uploaded_file.name}"
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        
        # Add to vault
        st.session_state.document_vault.add_document(
            doc_name, doc_type, validity.isoformat(), file_path
        )
        st.success(f"✅ Document '{doc_name}' added to vault!")
    
    # Save resources
    if st.button("💾 Save Company Profile"):
        st.session_state.company_resources = {
            "workers": workers,
            "engineers": engineers,
            "vehicles": vehicles,
            "current_projects": projects
        }
        st.success("Company profile updated!")

elif tab == "📊 Tender Evaluation":
    st.header("📊 Comprehensive Tender Evaluation")

    if not st.session_state.analysis_results:
        st.warning("⚠️ No tender analysis available. Please analyze tenders first.")
        st.stop()

    avk5_data = json.load(open("../data/avk5_standards.json", "r", encoding="utf-8"))
    compliance = st.session_state.document_vault
    profitability = ProfitabilityAnalyzer(AVK5Estimator())

    tender_options = {r["tender_id"]: r["title"] for r in st.session_state.analysis_results}
    selected_tender = st.selectbox("Select tender for evaluation:", options=list(tender_options.keys()), format_func=lambda x: f"{tender_options[x][:50]}...")

    if not selected_tender:
        st.stop()

    tender_data = next(r for r in st.session_state.analysis_results if r["tender_id"] == selected_tender)

    # ---------------------- 📄 Document Compliance ----------------------
    with st.expander("📄 Document Compliance Check", expanded=True):
        if tender_data.get("required_documents"):
            doc_report = compliance.check_compliance(tender_data["required_documents"])
            col1, col2 = st.columns([1, 3])
            col1.metric("Compliance Score", f"{doc_report['compliance_score']*100:.1f}%")
            col2.metric("Status", "✅ Compliant" if doc_report["is_compliant"] else "❌ Non-Compliant")

            st.subheader("Document Status")
            for doc in tender_data["required_documents"]:
                status = "✅ Available" if doc in doc_report["available_documents"] else "❌ Missing"
                st.write(f"{status} - {doc}")
        else:
            st.info("No document requirements specified in this tender.")

    # ---------------------- 🧱 AVK5 Material Cost Estimation ----------------------
    st.subheader("🧱 Add Custom Materials for AVK5 Estimation")

    # Prepare dropdown materials
    material_options = []
    category_map = {}
    for cat, items in avk5_data.items():
        if isinstance(items, dict):
            for spec in items:
                material_options.append(f"{spec} ({cat})")
                category_map[spec] = cat

    if "custom_materials" not in st.session_state:
        st.session_state.custom_materials = []

    selected_material = st.selectbox("Select Material", material_options)
    mat_spec = selected_material.split(" (")[0]
    mat_cat = category_map[mat_spec]
    unit = avk5_data[mat_cat][mat_spec]["unit"]
    default_price = avk5_data[mat_cat][mat_spec]["price"]

    with st.form("material_form"):
        col1, col2, col3 = st.columns(3)
        qty = col1.number_input("Quantity", min_value=0.0, step=0.1, key="mat_qty")
        price = col2.number_input("Unit Price (UAH)", value=float(default_price), step=100.0, key="mat_price")
        col3.markdown(f"💡 Unit: **{unit}**")
        if st.form_submit_button("➕ Add Material"):
            if qty > 0:
                st.session_state.custom_materials.append({
                    "category": mat_cat,
                    "specification": mat_spec,
                    "quantity": qty,
                    "unit_price": price,
                    "total": qty * price
                })
                st.success(f"✅ {mat_spec} added!")

    if st.session_state.custom_materials:
        st.subheader("📦 Current Material Inputs")
        total = 0
        for m in st.session_state.custom_materials:
            st.write(f"- {m['specification']} ({m['category']}): {m['quantity']} × {m['unit_price']} = {m['total']:.2f} UAH")
            total += m["total"]
        st.markdown(f"### 💰 Total Material Cost: **{total:,.2f} UAH**")

        # Export to Excel
        if st.button("📥 Download AVK5 Material Costs (Excel)"):
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet(title="AVK5 Custom Materials")
            else:
                ws.title = "AVK5 Custom Materials"
            headers = ["Category", "Specification", "Quantity", "Unit Price", "Total"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                ws.column_dimensions[get_column_letter(col_num)].width = 20
            for i, m in enumerate(st.session_state.custom_materials, 2):
                ws.cell(row=i, column=1, value=m["category"])
                ws.cell(row=i, column=2, value=m["specification"])
                ws.cell(row=i, column=3, value=m["quantity"])
                ws.cell(row=i, column=4, value=m["unit_price"])
                ws.cell(row=i, column=5, value=m["total"])
            ws.cell(row=len(st.session_state.custom_materials) + 2, column=4, value="Total:")
            ws.cell(row=len(st.session_state.custom_materials) + 2, column=5, value=total)

            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            st.download_button("📥 Download AVK5 Estimate", data=excel_buffer, file_name=f"avk5_materials_{selected_tender}.xlsx")

    # ---------------------- 💰 Profitability Analysis ----------------------
    with st.expander("💰 Profitability Analysis", expanded=True):
        def extract_resources(text):
            import re
            resources = {"workers": 0, "engineers": 0, "vehicles": 0}
            patterns = {
                "workers": r"(\d+)\s*(workers|laborers|people)",
                "engineers": r"(\d+)\s*(engineers)",
                "vehicles": r"(\d+)\s*(vehicles|trucks|cars)"
            }
            for key, pattern in patterns.items():
                found = re.search(pattern, text, re.IGNORECASE)
                if found:
                    resources[key] = int(found.group(1))
            return resources

        def estimate_complexity(text):
            text = text.lower()
            if any(w in text for w in ["automation", "bim", "hvac", "deep foundation"]): return 8
            if any(w in text for w in ["roof", "paving", "electrical"]): return 5
            if any(w in text for w in ["painting", "doors"]): return 3
            return 4

        def safe_budget(budget_raw):
            try:
                return float(budget_raw.split()[0].replace(",", ""))
            except:
                return 0.0

        auto_resource_req = extract_resources(tender_data.get("resource_requirements", ""))
        auto_complexity = estimate_complexity(tender_data.get("technical_specs", ""))
        auto_budget = safe_budget(tender_data.get("budget", "0"))
        # Calculate total cost from AVK5 custom materials if available
        custom_materials = st.session_state.get("custom_materials", [])
        estimated_cost = 0

        if custom_materials:
            estimated_cost = sum(m["total"] for m in custom_materials)
            st.markdown(f"💸 **Estimated Cost from AVK5 Inputs**: `{estimated_cost:,.2f} UAH`")

        tender = {
            "title": tender_data.get("title", ""),
            "budget": auto_budget,
            "resource_requirements": auto_resource_req,
            "estimated_cost": estimated_cost,
            "timeline": {
                "duration_days": 90,
                "start_date": "2025-09-01"
            },
            "complexity": auto_complexity,
            "payment_terms": tender_data.get("payment_terms", "standard").lower(),
            "has_penalties": False,
            "competitors": 3,
            "required_docs": tender_data.get("required_documents", [])
        }

        company = st.session_state.company_resources
        analysis = profitability.analyze_tender(tender, company)

        col1, col2, col3 = st.columns(3)
        col1.metric("ROI Score", f"{analysis['roi_score']:.1f}/100")
        col2.metric("Profit Margin", f"{analysis['profit_margin']*100:.1f}%")
        col3.metric("Recommendation", analysis["recommendation"])

        st.subheader("Resource Gap Analysis")
        st.subheader("📈 Financial Breakdown")
        st.markdown(f"- **Tender Value**: {analysis['tender_value']:,.2f} UAH")
        st.markdown(f"- **Estimated Cost**: {analysis['estimated_cost']:,.2f} UAH")
        st.markdown(f"- **Gross Profit**: {analysis['gross_profit']:,.2f} UAH")

        st.subheader("📦 Cost Breakdown")
        for cat, val in analysis.get("cost_breakdown", {}).items():
            if isinstance(val, (int, float)):
                st.write(f"- {cat}: {val:,.2f} UAH")
            elif isinstance(val, list):
                st.write(f"- {cat}:")
                for item in val:
                    st.write(f"    {item}")
            else:
                st.write(f"- {cat}: {val}")

        st.subheader("⏱️ Timeline Feasibility")
        st.info(f"{analysis.get('timeline_feasibility', 'N/A')}")

        st.subheader("⚠️ Risk Factors")
        for factor in analysis.get("risk_factors", []):
            st.warning(f"- {factor}")

        if analysis["resource_gap"].get("gap_analysis"):
            for resource, data in analysis["resource_gap"]["gap_analysis"].items():
                gap_percent = max(0, min(1, data["gap_percent"]))
                st.progress(1 - gap_percent, text=f"{resource}: {data['available']}/{data['required']} ({data['gap']} gap)")
        else:
            st.info("No resource requirements specified.")

        # Export Evaluation to Excel
        if st.button("📥 Download Evaluation Report (Excel)"):
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet(title="Tender Evaluation")
            else:
                ws.title = "Tender Evaluation"
                
            ws.append(["Tender Title", tender["title"]])
            ws.append(["Budget", tender["budget"]])
            ws.append(["Duration (days)", tender["timeline"]["duration_days"]])
            ws.append(["Complexity", tender["complexity"]])
            ws.append(["Payment Terms", tender["payment_terms"]])
            ws.append(["Has Penalties", tender["has_penalties"]])
            ws.append(["Competitors", tender["competitors"]])
            ws.append(["ROI Score", analysis["roi_score"]])
            ws.append(["Profit Margin", analysis["profit_margin"]])
            ws.append(["Recommendation", analysis["recommendation"]])
            ws.append([])

            # Financial Breakdown
            ws.append(["📈 Financial Breakdown"])
            ws.append(["Tender Value", analysis["tender_value"]])
            ws.append(["Estimated Cost", analysis["estimated_cost"]])
            ws.append(["Gross Profit", analysis["gross_profit"]])
            ws.append([])

            # Cost Breakdown
            ws.append(["📦 Cost Breakdown"])
            ws.append(["Category", "Amount (UAH)"])
            for cat, val in analysis.get("cost_breakdown", {}).items():
                ws.append([cat, val])
            ws.append([])

            # Timeline Feasibility
            ws.append(["⏱️ Timeline Feasibility", analysis.get("timeline_feasibility", "N/A")])
            ws.append([])

            # Risk Factors
            ws.append(["⚠️ Risk Factors"])
            for risk in analysis.get("risk_factors", []):
                ws.append([risk])
            ws.append([])

            # Resource Gaps
            ws.append(["Resource", "Required", "Available", "Gap", "Gap %"])
            for resource, data in analysis["resource_gap"]["gap_analysis"].items():
                ws.append([
                    resource,
                    data["required"],
                    data["available"],
                    data["gap"],
                    f"{data['gap_percent']*100:.1f}%"
                ])

            # Export
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            st.download_button(
                label="📊 Download Evaluation Report (Excel)",
                data=buffer,
                file_name=f"tender_evaluation_{selected_tender}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
