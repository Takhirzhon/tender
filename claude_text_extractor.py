import os
import json
import time
import anthropic
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

client = anthropic.Anthropic(
    api_key=os.getenv("CLAUDE_API_KEY")
)

TEXT_DIR = "tenders/text"
OUTPUT_EXCEL = "tenders/claude_extracted.xlsx" 
MAX_FILES = 20 
MAX_TOKENS = 8000 

COLUMNS = ["Title", "Issuer", "Deadline", "Budget", "Location", "Project Type", "Filename"]
data = []

def ask_claude(text):
    prompt = f"""
Extract the following fields from the tender text below:
1. Title or Project Name
2. Issuer or Client
3. Submission Deadline
4. Estimated Budget
5. Location (City, Region, or Country)
6. Project Type or Scope of Work

Return the result strictly in the following JSON format:
{{
  "title": "...",
  "issuer": "...",
  "deadline": "...",
  "budget": "...",
  "location": "...",
  "project_type": "..."
}}

Tender Text:
\"\"\"
{text}
\"\"\"
"""
    response = client.messages.create(
        model="claude-3-5-sonnet-20241022",
        max_tokens=1024,
        temperature=0.0,
        system="You are a helpful assistant that extracts structured information from public procurement tender documents.",
        messages=[{"role": "user", "content": prompt}]
    )
    
    if response.content:
        return ''.join(block.text for block in response.content if hasattr(block, 'text')) #type: ignore
    return ""

def format_excel(file_path):
    """Apply formatting to the Excel output"""
    wb = Workbook()
    ws = wb.active
    
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, column_title in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)  # type: ignore
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    col_widths = [40, 40, 20, 20, 30, 30, 30] 
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width  # type: ignore
    
    return wb, ws

print("⏳ Starting tender extraction with Claude...")
start_time = time.time()

wb, ws = format_excel(OUTPUT_EXCEL)
row_counter = 2 

processed_count = 0
for filename in os.listdir(TEXT_DIR):
    if not filename.endswith(".txt") or processed_count >= MAX_FILES:
        continue

    filepath = os.path.join(TEXT_DIR, filename)
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()

        print(f"🔍 Processing ({processed_count+1}/{MAX_FILES}): {filename}")
        result = ask_claude(content[:MAX_TOKENS])
        
        try:
            parsed = json.loads(result)
        except json.JSONDecodeError:
            import re
            json_match = re.search(r'\{[\s\S]*\}', result)
            if json_match:
                try:
                    parsed = json.loads(json_match.group())
                except:
                    parsed = {}
            else:
                parsed = {}
        
        row_data = {
            "Title": parsed.get("title", "Not extracted"),
            "Issuer": parsed.get("issuer", "Not extracted"),
            "Deadline": parsed.get("deadline", "Not specified"),
            "Budget": parsed.get("budget", "Not specified"),
            "Location": parsed.get("location", "Not specified"),
            "Project Type": parsed.get("project_type", "Not specified"),
            "Filename": filename
        }
        
        for col_num, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_counter, column=col_num, value=row_data[col_name])  # type: ignore
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row_counter += 1
        processed_count += 1
        time.sleep(1.2) 
        
    except Exception as e:
        print(f"❌ Error processing {filename}: {str(e)}")
        for col_num, col_name in enumerate(COLUMNS, 1):
            value = "ERROR" if col_name != "Filename" else filename
            ws.cell(row=row_counter, column=col_num, value=value)  # type: ignore
        row_counter += 1

try:
    for row in ws.iter_rows(min_row=2, max_row=row_counter-1, max_col=len(COLUMNS)):  # type: ignore
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    wb.save(OUTPUT_EXCEL)
    proc_time = time.time() - start_time
    
    print(f"\n✅ Successfully processed {processed_count} tenders")
    print(f"💾 Excel file saved to: {OUTPUT_EXCEL}")
    print(f"⏱️ Total processing time: {proc_time:.2f} seconds")
    print(f"⏳ Average time per tender: {proc_time/processed_count if processed_count else 0:.2f} seconds")
    
except Exception as e:
    print(f"❌ Failed to save Excel file: {str(e)}")

print("🏁 Extraction complete")